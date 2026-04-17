import openpyxl
PATH = "assets/floor_1_data.xlsx"
wb = openpyxl.load_workbook(PATH)
ws = wb['paper']
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
if 'Mark' not in headers:
    col = ws.max_column + 1
    ws.cell(1, col).value = 'Mark'
    headers.append('Mark')
mark_col = headers.index('Mark') + 1
name_col  = headers.index('Name') + 1
marks = { 'T.Slip': '1', 'Batch Record (Can)': 'C', 'Batch Record (Pouch)': 'C' }
for row in range(2, ws.max_row + 1):
    name = ws.cell(row, name_col).value
    if name in marks:
        ws.cell(row, mark_col).value = marks[name]
        print(f"Row {row} '{name}' → Mark={marks[name]}")
wb.save(PATH)
print("Saved.")
